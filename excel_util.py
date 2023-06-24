import openpyxl
import xlsxwriter


def read_from_excel(file_location, file_id):
    sh = openpyxl.load_workbook(file_location).active
    reports = []
    for row in range(1, sh.max_row + 1):
        if sh.cell(row=row, column=2).value is not None:
            datum_vreme = str(sh.cell(row=row, column=4).value).split(" ")
            mesto = None
            if sh.cell(row=row, column=15).value:
                mesto = sh.cell(row=row, column=15).value.strip()
            report = (
                int(sh.cell(row=row, column=3).value),
                datum_vreme[0],
                datum_vreme[1],
                sh.cell(row=row, column=5).value.strip(),
                sh.cell(row=row, column=6).value.strip(),
                sh.cell(row=row, column=7).value.strip(),
                sh.cell(row=row, column=8).value.strip(),
                sh.cell(row=row, column=9).value.strip(),
                sh.cell(row=row, column=10).value.strip(),
                sh.cell(row=row, column=11).value.strip(),
                float(sh.cell(row=row, column=12).value),
                float(sh.cell(row=row, column=13).value),
                float(sh.cell(row=row, column=14).value),
                mesto,
                file_id
            )
            reports.append(report)
    return reports


def export_to_excel(file_name, data):
    workbook = xlsxwriter.Workbook(file_name)
    worksheet = workbook.add_worksheet("Sheet1")

    header_format = workbook.add_format()
    header_format.set_bold(True)
    header_format.set_font_size(16)
    header_format.set_font_name('Times New Roman CE')

    cell_format = workbook.add_format()
    cell_format.set_font_name('Times New Roman CE')
    cell_format.set_font_size(10)
    cell_format.set_border()

    worksheet.write(1, 1, "IZVEŠTAJ", header_format)
    worksheet.write(3, 1, "broj", cell_format)
    worksheet.write(3, 2, "datum", cell_format)
    worksheet.write(3, 3, "Pošiljalac", cell_format)
    worksheet.write(3, 4, "Poručilac", cell_format)
    worksheet.write(3, 5, "Primalac", cell_format)
    worksheet.write(3, 6, "Artikal", cell_format)
    worksheet.write(3, 7, "Prevoznik", cell_format)
    worksheet.write(3, 8, "Registracija", cell_format)
    worksheet.write(3, 9, "Vozač", cell_format)
    worksheet.write(3, 10, "Bruto (kg)", cell_format)
    worksheet.write(3, 11, "Tara (kg)", cell_format)
    worksheet.write(3, 12, "Neto (kg)", cell_format)
    worksheet.write(3, 13, "Mesto", cell_format)

    row = 4
    col = 1
    for e in data:
        cnt = row - 3
        worksheet.write_number(row, 0, cnt, cell_format)

        for key, value in e.items():
            if key in ['datum']:
                continue
            elif key in ['vreme']:
                worksheet.write(row, col, e['datum'] + " " + e['vreme'], cell_format)
            elif key in ['bruto', 'tara', 'neto']:
                worksheet.write_number(row, col, int(value.replace(',', '')), cell_format)
            elif key in ['broj']:
                worksheet.write_number(row, col, value, cell_format)
            else:
                worksheet.write(row, col, value, cell_format)
            col += 1
        col = 1
        row += 1
    worksheet.write(row, 12, f"=SUM(M5:M{row})")

    workbook.close()

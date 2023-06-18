import openpyxl


def read_from_excel(file_location, file_id):
    sh = openpyxl.load_workbook(file_location).active
    reports = []
    for row in range(1, sh.max_row+1):
        if sh.cell(row=row, column=2).value is not None:
            datum_vreme = str(sh.cell(row=row, column=4).value).split(" ")
            report = (
                int(sh.cell(row=row, column=3).value),
                datum_vreme[0],
                datum_vreme[1],
                sh.cell(row=row, column=5).value.strip(),
                sh.cell(row=row, column=6).value.strip(),
                sh.cell(row=row, column=7).value.strip(),
                sh.cell(row=row, column=8).value.strip(),
                sh.cell(row=row, column=9).value.strip(),
                sh.cell(row=row, column=10).value.strip(),
                sh.cell(row=row, column=11).value.strip(),
                float(sh.cell(row=row, column=12).value),
                float(sh.cell(row=row, column=13).value),
                float(sh.cell(row=row, column=14).value),
                file_id
            )
            reports.append(report)
    return reports
